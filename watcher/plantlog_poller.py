r"""Phase 6.6 — Plantlog XLSX-export poller.

Hourly Task Scheduler job. Logs into plantlog via cookie auth, fetches two
memorized reports as XLSX, parses with openpyxl, upserts into Supabase.

Memorized reports being polled:
  7 — "Past 7days logs by user"     -> plantlog_log_records
  8 — "Log Record Latest ALL"       -> plantlog_latest_readings

Schedule (set up via install_plantlog_poller_task.ps1):
  Fires hourly 7:00am - 7:00pm daily.

Run manually:
    .\.venv\Scripts\python.exe plantlog_poller.py
"""
from __future__ import annotations

import io
import os
import re
import sys
from datetime import datetime, date, time as time_cls
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import openpyxl
import requests
from dotenv import load_dotenv

HERE = Path(__file__).resolve().parent
load_dotenv(HERE / ".env")

sys.path.insert(0, str(HERE))
from supabase_client import get_client  # noqa: E402
from plantlog_session import login, SessionError  # noqa: E402
from plantlog_building_attribution import attribute_and_persist  # noqa: E402

BASE_URL = os.environ.get("PLANTLOG_BASE_URL", "https://cwservices-bmrupark.plantlog.com").rstrip("/")
GQL_PATH = "/plantlog/api"
EASTERN = ZoneInfo("America/New_York")
UTC = ZoneInfo("UTC")

MEMO_LOG_RECORDS = 7
MEMO_LATEST_READINGS = 8

# Plantlog's "May 21, 2026" / "May 21, 2026 @ 09:39" — short-month + day + year.
DATE_FMT = "%b %d, %Y"
DATETIME_FMT = "%b %d, %Y @ %H:%M"


# ---------- HTTP helpers ----------

def fetch_xlsx(cookies: dict[str, str], memo_id: int) -> bytes:
    url = f"{BASE_URL}{GQL_PATH}/reports/memorized/{memo_id}/XLSX"
    headers = {
        "Accept": "*/*",
        "Referer": f"{BASE_URL}/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
    }
    resp = requests.get(url, headers=headers, cookies=cookies, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"memo {memo_id} XLSX HTTP {resp.status_code}: {resp.text[:300]}")
    if not resp.content.startswith(b"PK"):
        raise RuntimeError(f"memo {memo_id} XLSX response is not a valid xlsx (no PK magic)")
    return resp.content


def _open_workbook(data: bytes):
    # NOTE: read_only=True is broken on JasperReports-generated xlsx —
    # ws.iter_rows() returns malformed single-cell rows. The files are
    # small (<200KB) so the memory cost of full load is negligible.
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return wb


def _as_str(v: Any) -> str:
    """Stringify a cell, handling None / date / time / datetime / numeric."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime(DATETIME_FMT)
    if isinstance(v, date):
        return v.strftime(DATE_FMT)
    if isinstance(v, time_cls):
        return v.strftime("%H:%M")
    return str(v).strip()


# ---------- Datetime conversion ----------

def local_to_utc(dt: datetime) -> datetime:
    """Treat a naive datetime as America/New_York wall-clock, return UTC-aware."""
    return dt.replace(tzinfo=EASTERN).astimezone(UTC)


def combine_local_to_utc(d: date, t: time_cls) -> datetime:
    return local_to_utc(datetime.combine(d, t))


# ---------- Parsers ----------

LOG_RECORDS_HEADER = ("Date", "Time", "Group", "Log", "Activity")


def parse_log_records(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Memo 7: scan rows, track current user from section headers."""
    wb = _open_workbook(xlsx_bytes)
    ws = wb[wb.sheetnames[0]]
    rows_out: list[dict[str, Any]] = []
    current_user: str | None = None

    for row in ws.iter_rows(values_only=True):
        cells = [_as_str(c) for c in row]
        while len(cells) < 5:
            cells.append("")
        if not any(cells[:5]):
            continue

        c0, c1, c2, c3, c4 = cells[:5]

        # Column-header row
        if (c0, c1, c2, c3, c4) == LOG_RECORDS_HEADER:
            continue

        # Username section header: only c0 non-empty
        if c0 and not c1 and not c2 and not c3 and not c4:
            # Skip the report title + date-range pre-amble (first two rows)
            # by recognizing the patterns. Title contains words like "Past"
            # / "logs"; date range contains " @ " and " to ".
            low = c0.lower()
            if low.startswith("past ") or " to " in c0 or "@" in c0:
                continue
            current_user = c0
            continue

        # Data row: all 5 cells present and date-shaped
        if current_user and c0 and c1 and c3:
            try:
                performed_on = datetime.strptime(c0, DATE_FMT).date()
            except ValueError:
                continue
            try:
                hh, mm = c1.split(":")
                performed_at_local = time_cls(int(hh), int(mm))
            except (ValueError, IndexError):
                continue
            performed_at_utc = combine_local_to_utc(performed_on, performed_at_local)
            rows_out.append({
                "source_memo_id":     MEMO_LOG_RECORDS,
                "user_name":          current_user,
                "performed_on":       performed_on.isoformat(),
                "performed_at_local": performed_at_local.isoformat(),
                "performed_at_utc":   performed_at_utc.isoformat(),
                "group_name":         c2 or None,
                "log_name":           c3,
                "activity_name":      c4 or None,
                "raw_row":            {"date": c0, "time": c1, "group": c2, "log": c3, "activity": c4},
            })
    return rows_out


_LATEST_LABELS = {"Completed:", "User:", "Activity:", "Note:"}


def parse_latest_readings(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    """Memo 8: state machine over per-equipment blocks."""
    wb = _open_workbook(xlsx_bytes)
    ws = wb[wb.sheetnames[0]]
    rows_out: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    in_readings = False

    def emit():
        nonlocal current
        if not current:
            return
        completed_str = current.get("completed_str")
        log_name = current.get("log_name")
        if not completed_str or not log_name:
            current = None
            return
        try:
            completed_local = datetime.strptime(completed_str, DATETIME_FMT)
        except ValueError:
            current = None
            return
        completed_utc = local_to_utc(completed_local)
        rows_out.append({
            "source_memo_id":     MEMO_LATEST_READINGS,
            "log_name":           log_name,
            "completed_at_local": completed_local.isoformat(),
            "completed_at_utc":   completed_utc.isoformat(),
            "completed_by_user":  current.get("user") or None,
            "activity_name":      current.get("activity") or None,
            "note":               current.get("note") or None,
            "readings":           current.get("readings") or [],
        })
        current = None

    for row in ws.iter_rows(values_only=True):
        cells = [_as_str(c) for c in row]
        while len(cells) < 4:
            cells.append("")
        c0, c1, c2, c3 = cells[:4]

        # Fully empty row
        if not (c0 or c1 or c2 or c3):
            continue

        # Skip the preamble (first 2 rows): title + "All"
        if c0 == "Log Record Latest ALL" and not c1:
            continue
        if c0 == "All" and not c1 and not c2 and not c3:
            continue

        # Field labels inside the current block (order matters: check before
        # the "reading row" rule, because labels have c1 non-empty too).
        if c0 in _LATEST_LABELS:
            if current is not None:
                field = c0[:-1].lower()  # "completed" / "user" / "activity" / "note"
                if field == "completed":
                    current["completed_str"] = c1
                else:
                    current[field] = c1
            in_readings = False
            continue

        # Reading-section header
        if c0 == "Item" and c2 == "Unit" and c3 == "Value":
            in_readings = True
            continue

        # New equipment block: c0 set, c1/c2/c3 ALL empty. Check this BEFORE
        # the reading-row rule so that the next equipment doesn't get eaten as
        # a reading with blank unit + blank value.
        if c0 and not c1 and not c2 and not c3:
            emit()
            current = {"log_name": c0, "readings": []}
            in_readings = False
            continue

        # Reading row: c0 set + at least one of (c2, c3) — units may be blank
        # for boolean-style items like "Active Alarms: No" or status enums.
        if in_readings and c0 and current is not None:
            current["readings"].append({
                "item":  c0,
                "unit":  c2 or None,
                "value": c3,
            })

    emit()
    return rows_out


# ---------- Supabase writes ----------

def write_snapshot(client, kind: str, filename: str) -> str:
    snap = client.table("snapshots").insert({
        "kind":        kind,
        "taken_at":    datetime.now(UTC).isoformat(),
        "filename":    filename,
        "source_path": f"plantlog://memorized/{kind}",
    }).execute()
    return snap.data[0]["id"]


def write_ingestion_log(client, *, filename: str, kind: str, status: str,
                        rows: int, snapshot_id: str | None = None,
                        error_msg: str | None = None) -> None:
    try:
        client.table("ingestion_log").insert({
            "filename":    filename,
            "kind":        kind,
            "status":      status,
            "rows":        rows,
            "snapshot_id": snapshot_id,
            "error_msg":   error_msg[:4000] if error_msg else None,
        }).execute()
    except Exception as e:
        print(f"WARN: ingestion_log write failed: {e}", file=sys.stderr)


def _dedupe(rows: list[dict[str, Any]], key_fields: tuple[str, ...]) -> list[dict[str, Any]]:
    """Last-write-wins dedupe on the unique key. Postgres' ON CONFLICT can't
    handle dupes within a single upsert batch — they have to be collapsed
    client-side first."""
    seen: dict[tuple, dict[str, Any]] = {}
    for r in rows:
        key = tuple(r[f] for f in key_fields)
        seen[key] = r
    return list(seen.values())


def upsert_log_records(client, snapshot_id: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    rows = _dedupe(rows, ("user_name", "log_name", "performed_at_utc"))
    for r in rows:
        r["snapshot_id"] = snapshot_id
    CHUNK = 500
    written = 0
    for i in range(0, len(rows), CHUNK):
        client.table("plantlog_log_records").upsert(
            rows[i:i + CHUNK],
            on_conflict="user_name,log_name,performed_at_utc",
            returning="minimal",
        ).execute()
        written += len(rows[i:i + CHUNK])
    return written


def upsert_latest_readings(client, snapshot_id: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    rows = _dedupe(rows, ("log_name", "completed_at_utc"))
    for r in rows:
        r["snapshot_id"] = snapshot_id
    CHUNK = 500
    written = 0
    for i in range(0, len(rows), CHUNK):
        client.table("plantlog_latest_readings").upsert(
            rows[i:i + CHUNK],
            on_conflict="log_name,completed_at_utc",
            returning="minimal",
        ).execute()
        written += len(rows[i:i + CHUNK])
    return written


# ---------- Main ----------

def synthetic_filename(memo_id: int, now_local: datetime) -> str:
    return f"plantlog-memo{memo_id}-{now_local.strftime('%Y-%m-%d %H-%M')}.xlsx"


def main() -> int:
    now_local = datetime.now(EASTERN)
    print(f"[{now_local.isoformat()}] plantlog poller starting")

    try:
        cookies = login()
    except SessionError as e:
        msg = f"plantlog login failed: {e}"
        print(f"ERROR: {msg}", file=sys.stderr)
        write_ingestion_log(get_client(),
                            filename=synthetic_filename(0, now_local),
                            kind="plantlog_records", status="error",
                            rows=0, error_msg=msg)
        return 1

    client = get_client()
    overall_ok = True

    # --- Memorized 7: log records ---
    fn7 = synthetic_filename(MEMO_LOG_RECORDS, now_local)
    try:
        xlsx = fetch_xlsx(cookies, MEMO_LOG_RECORDS)
        rows = parse_log_records(xlsx)
        snapshot_id = write_snapshot(client, "plantlog_records", fn7)
        n = upsert_log_records(client, snapshot_id, rows)
        client.table("snapshots").update({"row_count": n}).eq("id", snapshot_id).execute()
        write_ingestion_log(client, filename=fn7, kind="plantlog_records",
                            status="ok", rows=n, snapshot_id=snapshot_id)
        print(f"[ok] memo 7 log_records: parsed {len(rows)} rows, upserted {n}")
    except Exception as e:
        overall_ok = False
        msg = f"memo 7 failed: {e}"
        print(f"ERROR: {msg}", file=sys.stderr)
        write_ingestion_log(client, filename=fn7, kind="plantlog_records",
                            status="error", rows=0, error_msg=msg)

    # --- Memorized 8: latest readings ---
    fn8 = synthetic_filename(MEMO_LATEST_READINGS, now_local)
    try:
        xlsx = fetch_xlsx(cookies, MEMO_LATEST_READINGS)
        rows = parse_latest_readings(xlsx)
        snapshot_id = write_snapshot(client, "plantlog_latest", fn8)
        n = upsert_latest_readings(client, snapshot_id, rows)
        client.table("snapshots").update({"row_count": n}).eq("id", snapshot_id).execute()
        write_ingestion_log(client, filename=fn8, kind="plantlog_latest",
                            status="ok", rows=n, snapshot_id=snapshot_id)
        print(f"[ok] memo 8 latest_readings: parsed {len(rows)} blocks, upserted {n}")
    except Exception as e:
        overall_ok = False
        msg = f"memo 8 failed: {e}"
        print(f"ERROR: {msg}", file=sys.stderr)
        write_ingestion_log(client, filename=fn8, kind="plantlog_latest",
                            status="error", rows=0, error_msg=msg)

    # Phase 6.7 — refresh building_inferred on the most-recent rows so the
    # dashboard sees attributed data. 14d covers the per-engineer breakdown
    # plus enough trail for the period-toggle UI. Failure here is non-fatal
    # — pollers already ingested; attribution can be re-run manually.
    try:
        diag = attribute_and_persist(days=14, client=client)
        print(f"[ok] building attribution: {diag}")
    except Exception as e:
        print(f"WARN: building attribution failed: {e}", file=sys.stderr)

    return 0 if overall_ok else 1


if __name__ == "__main__":
    sys.exit(main())
